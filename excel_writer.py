"""
Excel Writer Module
====================
Creates a professionally styled Excel file (QnA.xlsx) with three
language-specific sheets: English, Hindi, and Marathi.
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


# Color scheme for each language sheet tab
SHEET_COLORS = {
    "English": "1F4E79",   # Deep Blue
    "Hindi": "C55A11",     # Warm Orange
    "Marathi": "548235",   # Forest Green
}

# Header background colors matching the sheet theme
HEADER_FILLS = {
    "English": PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid"),
    "Hindi": PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid"),
    "Marathi": PatternFill(start_color="548235", end_color="548235", fill_type="solid"),
}

# Alternating row colors for readability
ROW_FILLS = {
    "even": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
    "odd": PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
}

# Border style
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

HEADER_BORDER = Border(
    left=Side(style="thin", color="FFFFFF"),
    right=Side(style="thin", color="FFFFFF"),
    top=Side(style="thin", color="FFFFFF"),
    bottom=Side(style="medium", color="000000"),
)


def _create_sheet(workbook: Workbook, sheet_name: str, qna_pairs: list[dict], is_first: bool = False):
    """
    Create and style a single worksheet with QnA data.
    
    Args:
        workbook: The openpyxl Workbook object.
        sheet_name: Name of the sheet (English, Hindi, or Marathi).
        qna_pairs: List of dicts with 'question' and 'answer' keys.
        is_first: If True, use the default active sheet instead of creating new.
    """
    if is_first:
        ws = workbook.active
        ws.title = sheet_name
    else:
        ws = workbook.create_sheet(title=sheet_name)
    
    # Set sheet tab color
    ws.sheet_properties.tabColor = SHEET_COLORS.get(sheet_name, "000000")
    
    # Header styling
    header_font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_fill = HEADER_FILLS.get(sheet_name, HEADER_FILLS["English"])
    
    # Write headers
    headers = ["Questions", "Answers"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = header_fill
        cell.border = HEADER_BORDER
    
    # Data styling
    data_font = Font(name="Calibri", size=11)
    data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Write QnA data
    for row_idx, pair in enumerate(qna_pairs, start=2):
        question = pair.get("question", pair.get("Question", ""))
        answer = pair.get("answer", pair.get("Answer", ""))
        
        q_cell = ws.cell(row=row_idx, column=1, value=question)
        a_cell = ws.cell(row=row_idx, column=2, value=answer)
        
        # Apply styling
        fill = ROW_FILLS["even"] if row_idx % 2 == 0 else ROW_FILLS["odd"]
        for cell in [q_cell, a_cell]:
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = THIN_BORDER
            cell.fill = fill
    
    # Set column widths
    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 65
    
    # Set row height for header
    ws.row_dimensions[1].height = 30
    
    # Freeze the header row
    ws.freeze_panes = "A2"


def create_excel(
    english_qna: list[dict],
    hindi_qna: list[dict],
    marathi_qna: list[dict],
    output_path: str = None,
) -> io.BytesIO:
    """
    Create a styled Excel file with three language-specific sheets.
    
    Args:
        english_qna: List of QnA dicts for English sheet.
        hindi_qna: List of QnA dicts for Hindi sheet.
        marathi_qna: List of QnA dicts for Marathi sheet.
        output_path: Optional file path to save the Excel file.
                     If None, returns a BytesIO buffer.
    
    Returns:
        BytesIO buffer containing the Excel file data.
    """
    wb = Workbook()
    
    # Create sheets in order: English, Hindi, Marathi
    _create_sheet(wb, "English", english_qna, is_first=True)
    _create_sheet(wb, "Hindi", hindi_qna, is_first=False)
    _create_sheet(wb, "Marathi", marathi_qna, is_first=False)
    
    # Save to file if path provided
    if output_path:
        wb.save(output_path)
    
    # Always return a BytesIO buffer for download
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer
